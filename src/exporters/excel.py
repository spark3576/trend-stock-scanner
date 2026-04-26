"""4시트 엑셀 산출물 생성기 (openpyxl).

시트 구성:
  00_summary    — 트리거 일자, 소스별 수집 건수, 스크리닝 통과 수
  10_trends     — 트렌드 키워드 raw
  20_screen1    — 1차 스크리닝 (지속 상승)
  30_screen2    — 2차 스크리닝 (2배 급증)
  40_momentum   — 종합 모멘텀 분석 + 근거
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.utils.logger import logger

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")


def _write_dataframe(ws, df: pd.DataFrame, start_row: int = 1) -> int:
    """DataFrame을 워크시트에 쓰고 마지막 사용 row를 반환."""
    if df.empty:
        ws.cell(row=start_row, column=1, value="(데이터 없음)")
        return start_row
    # 헤더
    for col_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=str(col))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # 데이터
    for r_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for c_idx, val in enumerate(row, 1):
            if pd.isna(val):
                val = ""
            elif isinstance(val, (pd.Timestamp,)):
                val = val.date().isoformat()
            ws.cell(row=r_idx, column=c_idx, value=val)
    # 컬럼 너비 자동
    for col_idx, col in enumerate(df.columns, 1):
        try:
            max_len = max(
                [len(str(col))] +
                [len(str(v)) for v in df.iloc[:200, col_idx - 1].astype(str).tolist()]
            )
        except Exception:
            max_len = 12
        ws.column_dimensions[get_column_letter(col_idx)].width = min(60, max(10, max_len + 2))
    return start_row + len(df)


def export_excel(
    out_path: Path,
    trigger_date: date,
    trends_df: pd.DataFrame,
    primary_df: pd.DataFrame,
    secondary_df: pd.DataFrame,
    momentum_df: pd.DataFrame,
) -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # 0. summary
    ws = wb.active
    ws.title = "00_summary"
    ws["A1"] = f"📈 Trend Stock Scanner — {trigger_date.isoformat()}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")

    ws["A3"] = "구분"
    ws["B3"] = "건수"
    for c in ["A3", "B3"]:
        ws[c].fill = HEADER_FILL
        ws[c].font = HEADER_FONT
    rows = [
        ("트렌드 키워드 (전체)", len(trends_df)),
        ("  - Google", int((trends_df["source"] == "google").sum()) if not trends_df.empty else 0),
        ("  - Naver",  int((trends_df["source"] == "naver").sum())  if not trends_df.empty else 0),
        ("  - Reddit", int((trends_df["source"] == "reddit").sum()) if not trends_df.empty else 0),
        ("  - Web",    int((trends_df["source"] == "web").sum())    if not trends_df.empty else 0),
        ("1차 스크리닝 통과 (지속 상승)", len(primary_df)),
        ("2차 스크리닝 통과 (2배 급증)", len(secondary_df)),
        ("모멘텀 종합 (트렌드+거래량)", len(momentum_df)),
        ("  - 🔴 강력주목", int((momentum_df["classification"] == "🔴 강력주목").sum()) if not momentum_df.empty else 0),
        ("  - 🟡 주목",     int((momentum_df["classification"] == "🟡 주목").sum())     if not momentum_df.empty else 0),
        ("  - ⚪ 관찰",     int((momentum_df["classification"] == "⚪ 관찰").sum())     if not momentum_df.empty else 0),
    ]
    for i, (k, v) in enumerate(rows, 4):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14

    ws.cell(row=len(rows) + 5, column=1,
            value="※ 본 결과는 정보 제공 목적이며 투자 권유가 아닙니다.")
    ws.cell(row=len(rows) + 5, column=1).font = Font(italic=True, color="888888")

    # 1. trends
    ws = wb.create_sheet("10_trends")
    _write_dataframe(ws, trends_df)

    # 2. primary screening
    ws = wb.create_sheet("20_screen1")
    _write_dataframe(ws, primary_df)

    # 3. secondary screening
    ws = wb.create_sheet("30_screen2")
    _write_dataframe(ws, secondary_df)

    # 4. momentum
    ws = wb.create_sheet("40_momentum")
    _write_dataframe(ws, momentum_df)

    wb.save(out_path)
    logger.info(f"[Excel] saved → {out_path}")
    return out_path
